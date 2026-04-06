import os
from pathlib import Path
from dotenv import load_dotenv
from supabase import create_client
from openai import OpenAI
from openpyxl import load_workbook

# ----------------------------
# CONFIG
# ----------------------------

BASE_DIR = Path(__file__).resolve().parent.parent
ENV_PATH = BASE_DIR / ".env"
EXCEL_DIR = BASE_DIR / "data" / "excel"

load_dotenv(ENV_PATH)

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")

if not SUPABASE_URL or not SUPABASE_KEY:
    raise ValueError("SUPABASE_URL ou SUPABASE_KEY manquant dans .env")

if not OPENAI_API_KEY:
    raise ValueError("OPENAI_API_KEY manquant dans .env")

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
openai_client = OpenAI(api_key=OPENAI_API_KEY)

# ----------------------------
# EMBEDDINGS
# ----------------------------

def get_embedding(text: str):
    response = openai_client.embeddings.create(
        model="text-embedding-3-small",
        input=text
    )
    return response.data[0].embedding

# ----------------------------
# EXTRACTION EXCEL
# ----------------------------

def extract_text_from_excel(file_path: Path):
    wb = load_workbook(filename=file_path, data_only=True)

    all_blocks = []

    for sheet in wb.worksheets:
        all_blocks.append(f"FEUILLE : {sheet.title}")

        for row in sheet.iter_rows(values_only=True):
            row_values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if row_values:
                line = " | ".join(row_values)
                all_blocks.append(line)

        all_blocks.append("")

    clean_text = "\n".join(all_blocks).strip()
    title = file_path.stem

    return title, clean_text

# ----------------------------
# NORMALISATION EN PARAGRAPHES
# ----------------------------

def normalize_paragraphs(text: str):
    raw_lines = text.splitlines()

    paragraphs = []
    current = []

    for line in raw_lines:
        stripped = line.strip()

        if not stripped:
            if current:
                paragraph = " ".join(current).strip()
                if paragraph:
                    paragraphs.append(paragraph)
                current = []
            continue

        current.append(stripped)

    if current:
        paragraph = " ".join(current).strip()
        if paragraph:
            paragraphs.append(paragraph)

    cleaned = []
    for p in paragraphs:
        p = " ".join(p.split())
        if len(p) >= 20:
            cleaned.append(p)

    return cleaned

# ----------------------------
# CHUNKING PROPRE
# ----------------------------

def split_chunks(text: str, max_size: int = 1500, overlap_paragraphs: int = 1):
    paragraphs = normalize_paragraphs(text)

    if not paragraphs:
        return []

    chunks = []
    current_chunk = []
    current_length = 0

    for paragraph in paragraphs:
        paragraph_length = len(paragraph)

        if paragraph_length > max_size:
            if current_chunk:
                chunks.append("\n\n".join(current_chunk).strip())
                current_chunk = []
                current_length = 0

            start = 0
            while start < paragraph_length:
                part = paragraph[start:start + max_size].strip()
                if part:
                    chunks.append(part)
                start += max_size
            continue

        if current_length + paragraph_length + 2 <= max_size:
            current_chunk.append(paragraph)
            current_length += paragraph_length + 2
        else:
            if current_chunk:
                chunks.append("\n\n".join(current_chunk).strip())

            overlap = current_chunk[-overlap_paragraphs:] if current_chunk else []
            current_chunk = overlap + [paragraph]
            current_length = sum(len(p) + 2 for p in current_chunk)

    if current_chunk:
        chunks.append("\n\n".join(current_chunk).strip())

    return chunks

# ----------------------------
# DB CHECK
# ----------------------------

def get_existing_document_by_file_name(file_name: str):
    response = (
        supabase
        .table("documents")
        .select("id, title, file_name")
        .eq("file_name", file_name)
        .limit(1)
        .execute()
    )

    if response.data and len(response.data) > 0:
        return response.data[0]

    return None

# ----------------------------
# INSERT CHUNKS
# ----------------------------

def insert_chunks(document_id: int, content_text: str):
    chunks = split_chunks(content_text)

    rows = []
    for index, chunk in enumerate(chunks):
        embedding = get_embedding(chunk)

        rows.append({
            "document_id": document_id,
            "chunk_index": index,
            "chunk_text": chunk,
            "embedding": embedding
        })

    if rows:
        supabase.table("document_chunks").insert(rows).execute()

    return len(rows)

# ----------------------------
# IMPORT EXCEL
# ----------------------------

def import_excel_to_supabase(file_path: Path, category: str = "excel"):
    if not file_path.exists():
        raise FileNotFoundError(f"Fichier Excel introuvable : {file_path}")

    existing_document = get_existing_document_by_file_name(file_path.name)

    if existing_document:
        return {
            "status": "already_exists",
            "message": "Ce fichier Excel existe déjà dans la base.",
            "document_id": existing_document["id"],
            "title": existing_document["title"],
            "file_name": existing_document["file_name"]
        }

    title, content_text = extract_text_from_excel(file_path)

    if not content_text.strip():
        return {
            "status": "empty",
            "message": "Aucun texte extractible trouvé dans ce fichier Excel.",
            "file_name": file_path.name
        }

    document_data = {
        "source_type": "excel",
        "title": title,
        "source_url": None,
        "file_name": file_path.name,
        "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "category": category,
        "status": "active",
        "content_text": content_text,
    }

    document_result = supabase.table("documents").insert(document_data).execute()

    inserted_document = document_result.data[0]
    document_id = inserted_document["id"]

    chunk_count = insert_chunks(document_id, content_text)

    return {
        "status": "inserted",
        "document_id": document_id,
        "title": title,
        "file_name": file_path.name,
        "chunk_count": chunk_count
    }

# ----------------------------
# MAIN
# ----------------------------

if __name__ == "__main__":
    files = sorted(EXCEL_DIR.rglob("*.xlsx"))

    print(f"Nombre de fichiers Excel trouvés : {len(files)}")

    for i, file in enumerate(files, 1):
        print("-" * 60)
        print(f"[{i}/{len(files)}] {file}")

        try:
            result = import_excel_to_supabase(file, category=file.parent.name)
            print(result)
        except Exception as e:
            print("ERREUR :", str(e))